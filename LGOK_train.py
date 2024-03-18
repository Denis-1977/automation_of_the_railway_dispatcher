import PySimpleGUI as Sg
from collections import defaultdict
from openpyxl import load_workbook
from collections import Counter
import re
import xlsxwriter

layout = [
        [Sg.Text('Введите количество вагонов в поезде:')],
        [Sg.Input()],
        [Sg.Button('Сформировать отчет')]
]
window = Sg.Window('Подход на ЛГОК', layout)
kolvag = 0
while True:
    event, values = window.read()
    if event == Sg.WINDOW_CLOSED:
        break
    elif event == 'Сформировать отчет':
        kolvag = int(values[0])

    window.close()

layout = [[Sg.Text("Отчет сформирован!")], [Sg.Button("ОК")]]
window = Sg.Window("Формирование отчета...", layout, finalize=True, size=(350, 100))

wb = load_workbook('Подход ЛГОК.xlsx')

sheet = wb.active

max_rows = sheet.max_row

index_dict = {}

owner_counts = defaultdict(lambda: defaultdict(int))
owner_counts_inn = defaultdict(lambda: defaultdict(int))

codes = {
        'НПК': '1',
        'инн': '2',
        'БП': '3',
        'ТТ': '4',
        'Нвтр': '5',
        'КФТ': '6',
        'ФГК': '7',
        'ПГК': '8'
    }
def owner_key(owner):
    return int(codes.get(owner, ''))

codes_total = {
        'НПК+договорные': '1',
        'Нвтр': '2',
        'КФТ': '3',
        'ФГК': '4',
        'ПГК': '5'
    }
def owner_key_total(owner):
    return codes_total.get(owner, float('inf'))

for i in range(5, max_rows+1):
    nambervag = sheet.cell(row=i, column=1).value

    owner = sheet.cell(row=i, column=2).value
    if owner is not None and 'нпк (иннов)' in owner.lower():
        owner = 'инн'
    elif owner is not None and 'нпк' in owner.lower():
        owner = 'НПК'
    elif owner is not None and 'новотранс' in owner.lower():
        owner = 'Нвтр'
    elif owner is not None and 'блисспро' in owner.lower():
        owner = 'БП'
    elif owner is not None and 'кфт' in owner.lower():
        owner = 'КФТ'
    elif owner is not None and 'транспортные технологии' in owner.lower():
        owner = 'ТТ'
    elif owner is not None and 'фгк' in owner.lower():
        owner = 'ФГК'
    elif owner is not None and 'пгк' in owner.lower():
        owner = 'ПГК'
    else:
        continue

    vagon = sheet.cell(row=i, column=3).value
    if vagon != 'Полувагоны (60)':
        continue

    weight = sheet.cell(row=i, column=23).value

    station = sheet.cell(row=i, column=32).value
    station = re.sub(r"\s\([^()]+\)", "", station)
    station = ' '.join(part.capitalize() for part in station.replace('-', ' ').split())
    if station == 'Старый Оскол':
        station = '   Ст.Оскол'
    elif station == 'Губкин':
        station = '  Губкин'
    elif station == 'Лебеди':
        station = '  Лебеди'
    elif station == 'Котел':
        station = '  Котел'
    elif station == 'Валуйки':
        station = ' Валуйки'
    elif station == 'Лихая':
        station = ' Лихая'
    elif station == 'Батайск':
        station = 'Батайск'
    elif station == 'Стойленская':
        continue

    road = sheet.cell(row=i, column=33).value
    if road == 'ГОРЬКОВСКАЯ (24)':
        road = 'Горьк.ж.д.'
    elif road == 'КУЙБЫШЕВСКАЯ (63)':
        road = 'Кбш.ж.д.'
    elif road == 'МОСКОВСКАЯ (17)':
        road = 'Мск.ж.д.'
    elif road == 'ПРИВОЛЖСКАЯ (61)':
        road = 'Прив.ж.д.'
    elif road == 'СЕВЕРНАЯ (28)':
        road = 'Сев.ж.д.'
    elif road == 'СЕВЕРО-КАВКАЗСКАЯ (51)':
        road = ' Скв.ж.д.'
    elif road == 'ЮГО-ВОСТОЧНАЯ (58)':
        road = ''
    elif road == 'ЮЖНО-УРАЛЬСКАЯ (80)':
        road = 'ЮУр.ж.д.'
    else:
        continue

    data_op = sheet.cell(row=i, column=36).value

    operation = sheet.cell(row=i, column=35).value

    index_r = sheet.cell(row=i, column=41).value
    match = re.search(r"(\d+)\s+(\d+)\s+(\d+)", index_r)
    if match:
        index_parts = match.groups()
        index_r = "{}-{}-{}".format(index_parts[0][:4], index_parts[1], index_parts[2][:4])
    if index_r == '0000-000-0000' or index_r.endswith('-5210') or index_r.endswith('-5209'):
        continue

    workbook = xlsxwriter.Workbook('ЛГОК.xlsx')
    worksheet = workbook.add_worksheet('Подход на Стойленскую')
    worksheet1 = workbook.add_worksheet('Брошенные поезда')
    worksheet2 = workbook.add_worksheet('Прямые поезда')

    row = 0
    row1 = 0
    row2 = 0

    index_dict.setdefault(index_r, []).append((station, road, operation, owner, nambervag, weight, data_op))
    owner_counts[index_r][owner] += 1
    owner_counts_inn[index_r][owner] += 1

    total_loaded_count_bros = {"НПК": 0, "инн": 0, "БП": 0, "ТТ": 0, "Нвтр": 0, "КФТ": 0, "ФГК": 0, "ПГК": 0}
    total_empty_count_bros = {"НПК": 0, "инн": 0, "БП": 0, "ТТ": 0, "Нвтр": 0, "КФТ": 0, "ФГК": 0, "ПГК": 0}

    total_loaded_count_direct = {"НПК": 0, "инн": 0, "БП": 0, "ТТ": 0, "Нвтр": 0, "КФТ": 0, "ФГК": 0, "ПГК": 0}
    total_empty_count_direct = {"НПК": 0, "инн": 0, "БП": 0, "ТТ": 0, "Нвтр": 0, "КФТ": 0, "ФГК": 0, "ПГК": 0}

    total_loaded_count_not_all = {"НПК": 0, "инн": 0, "БП": 0, "ТТ": 0, "Нвтр": 0, "КФТ": 0, "ФГК": 0, "ПГК": 0}
    total_empty_count_not_all = {"НПК": 0, "инн": 0, "БП": 0, "ТТ": 0, "Нвтр": 0, "КФТ": 0, "ФГК": 0, "ПГК": 0}

    total_loaded_count_all = {"НПК": 0, "инн": 0, "БП": 0, "ТТ": 0, "Нвтр": 0, "КФТ": 0, "ФГК": 0, "ПГК": 0}
    total_empty_count_all = {"НПК": 0, "инн": 0, "БП": 0, "ТТ": 0, "Нвтр": 0, "КФТ": 0, "ФГК": 0, "ПГК": 0}


    sorted_values = sorted(index_dict.items(), key=lambda x: (x[1][0][1], x[1][0][0]))

    for index_r, values_list in sorted_values:
        i = len(values_list)
        if i < kolvag:
            continue

        owners_loaded_count = {"НПК": 0, "инн": 0, "БП": 0, "ТТ": 0, "Нвтр": 0, "КФТ": 0, "ФГК": 0, "ПГК": 0}
        owners_empty_count = {"НПК": 0, "инн": 0, "БП": 0, "ТТ": 0, "Нвтр": 0, "КФТ": 0, "ФГК": 0, "ПГК": 0}

        npk_numbers = []
        other_numbers = []

        for station, road, operation, owner, nambervag, weight, data_op in values_list:

            # условие (итоговая сумма по собственникам)! все прямые, брошенные поезда на Ю-Вост.ж.д.
            if road == '' and index_r.endswith('-4384') and 'БРОС' in operation:
                if weight > 0:
                    total_loaded_count_bros[owner] += 1
                else:
                    total_empty_count_bros[owner] += 1

                owners_strings_set_bros = set()
                for owner in total_loaded_count_bros:
                    if total_loaded_count_bros[owner] > 0 or total_empty_count_bros[owner] > 0:

                        sum_4_loaded_owners_bros = sum([total_loaded_count_bros[owner] for owner in ["НПК", "инн", "БП", "ТТ"]])
                        sum_4_empty_owners_bros = sum([total_empty_count_bros[owner] for owner in ["НПК", "инн", "БП", "ТТ"]])
                        if owner in ["НПК", "инн", "БП", "ТТ"]:
                            if ("инн" and "НПК" and "БП" and "ТТ") in ["НПК", "инн", "БП", "ТТ"]:
                                owner = "НПК+договорные"

                            totals_bros = f"{sum_4_loaded_owners_bros}/{sum_4_empty_owners_bros}" if sum_4_loaded_owners_bros > 0 else f"{sum_4_empty_owners_bros}"
                        else:
                            totals_bros = f"{total_loaded_count_bros[owner]}/{total_empty_count_bros[owner]}" if total_loaded_count_bros[owner] > 0 else f"{total_empty_count_bros[owner]}"

                        total_string_bros = f"{owner}: {totals_bros}"
                        owners_strings_set_bros.add(total_string_bros)
                        unique_owner_strings_bros = list(owners_strings_set_bros)
                        sorted_owner_strings_bros = sorted(unique_owner_strings_bros, key=lambda x: owner_key_total(x.split(':')[0]))
                        total_string_bros = ', '.join(sorted_owner_strings_bros)
                        string_bros_train = f"{total_string_bros}"

            # условие (итоговая сумма по собственникам)! все прямые, НЕ брошенные поезда на Ю-Вост.ж.д.
            if road == '' and index_r.endswith('-4384') and not 'БРОС' in operation:
                if weight > 0:
                    total_loaded_count_direct[owner] += 1
                else:
                    total_empty_count_direct[owner] += 1

                owners_strings_set_direct = set()
                for owner in total_loaded_count_direct:
                    if total_loaded_count_direct[owner] > 0 or total_empty_count_direct[owner] > 0:

                        sum_4_loaded_owners_direct = sum([total_loaded_count_direct[owner] for owner in ["НПК", "инн", "БП", "ТТ"]])
                        sum_4_empty_owners_direct = sum([total_empty_count_direct[owner] for owner in ["НПК", "инн", "БП", "ТТ"]])
                        if owner in ["НПК", "инн", "БП", "ТТ"]:
                            if ("инн" and "НПК" and "БП" and "ТТ") in ["НПК", "инн", "БП", "ТТ"]:
                                owner = "НПК+договорные"

                            totals_direct = f"{sum_4_loaded_owners_direct}/{sum_4_empty_owners_direct}" if sum_4_loaded_owners_direct > 0 else f"{sum_4_empty_owners_direct}"
                        else:
                            totals_direct = f"{total_loaded_count_direct[owner]}/{total_empty_count_direct[owner]}" if total_loaded_count_direct[owner] > 0 else f"{total_empty_count_direct[owner]}"

                        total_string_direct = f"{owner}: {totals_direct}"
                        owners_strings_set_direct.add(total_string_direct)
                        unique_owner_strings_direct = list(owners_strings_set_direct)
                        sorted_owner_strings_direct = sorted(unique_owner_strings_direct, key=lambda x: owner_key_total(x.split(':')[0]))
                        total_string_direct = ', '.join(sorted_owner_strings_direct)
                        string_direct_train = f"{total_string_direct}"

            # условие (итоговая сумма по собственникам)! все поезда по дорогам (не прямые и не Ю-Вост.ж.д.)
            if not (road == '' and index_r.endswith('-4384')):
                if weight > 0:
                    total_loaded_count_not_all[owner] += 1
                else:
                    total_empty_count_not_all[owner] += 1

                owners_strings_set_not_all = set()
                for owner in total_loaded_count_not_all:
                    if total_loaded_count_not_all[owner] > 0 or total_empty_count_not_all[owner] > 0:

                        sum_4_loaded_owners_not_all = sum([total_loaded_count_not_all[owner] for owner in ["НПК", "инн", "БП", "ТТ"]])
                        sum_4_empty_owners_not_all = sum([total_empty_count_not_all[owner] for owner in ["НПК", "инн", "БП", "ТТ"]])
                        if owner in ["НПК", "инн", "БП", "ТТ"]:
                            if ("инн" and "НПК" and "БП" and "ТТ") in ["НПК", "инн", "БП", "ТТ"]:
                                owner = "НПК+договорные"

                            totals_not_all = f"{sum_4_loaded_owners_not_all}/{sum_4_empty_owners_not_all}" if sum_4_loaded_owners_not_all > 0 else f"{sum_4_empty_owners_not_all}"
                        else:
                            totals_not_all = f"{total_loaded_count_not_all[owner]}/{total_empty_count_not_all[owner]}" if total_loaded_count_not_all[owner] > 0 else f"{total_empty_count_not_all[owner]}"

                        total_string_not_all = f"{owner}: {totals_not_all}"
                        owners_strings_set_not_all.add(total_string_not_all)
                        unique_owner_strings_not_all = list(owners_strings_set_not_all)
                        sorted_owner_strings_not_all = sorted(unique_owner_strings_not_all, key=lambda x: owner_key_total(x.split(':')[0]))
                        total_string_not_all = ', '.join(sorted_owner_strings_not_all)
                        string_not_all_train = f"{total_string_not_all}"

        for station, road, operation, owner, nambervag, weight, data_op in values_list:

            if owner == 'НПК' or owner == 'инн':
                npk_numbers.append(nambervag)
                npk_nambervag = nambervag
            else:
                other_numbers.append(nambervag)
                other_nambervag = nambervag

            if weight > 0:
                owners_loaded_count[owner] += 1
            else:
                owners_empty_count[owner] += 1

            owners_strings_set = set()

            # итоговая сумма по собственникам (ВСЕГО)
            if weight > 0:
                total_loaded_count_all[owner] += 1
            else:
                total_empty_count_all[owner] += 1

            owners_strings_set_all = set()
            for owner in total_loaded_count_all:
                if total_loaded_count_all[owner] > 0 or total_empty_count_all[owner] > 0:

                    sum_4_loaded_owners_all = sum([total_loaded_count_all[owner] for owner in ["НПК", "инн",  "БП", "ТТ"]])
                    sum_4_empty_owners_all = sum([total_empty_count_all[owner] for owner in ["НПК", "инн",  "БП", "ТТ"]])
                    if owner in ["НПК", "инн", "БП", "ТТ"]:
                        if ("инн" and "НПК" and "БП" and "ТТ") in ["НПК", "инн",  "БП", "ТТ"]:
                            owner = "НПК+договорные"

                        totals_all = f"{sum_4_loaded_owners_all}/{sum_4_empty_owners_all}" if sum_4_loaded_owners_all > 0 else f"{sum_4_empty_owners_all}"
                    else:
                        totals_all = f"{total_loaded_count_all[owner]}/{total_empty_count_all[owner]}" if total_loaded_count_all[owner] > 0 else f"{total_empty_count_all[owner]}"

                    total_string_all = f"{owner}: {totals_all}"
                    owners_strings_set_all.add(total_string_all)
                    unique_owner_strings_all = list(owners_strings_set_all)
                    sorted_owner_strings_all = sorted(unique_owner_strings_all, key=lambda x: owner_key_total(x.split(':')[0]))
                    total_string_all = ', '.join(sorted_owner_strings_all)
                    string_all_train = f"{total_string_all}"

            # формирование отчета по собственникам в каждом поезде
            for owner in sorted(owners_loaded_count, key=owner_key):

                loaded_count = owners_loaded_count[owner]
                empty_count = owners_empty_count[owner]
                if loaded_count > 0 or empty_count > 0:

                    sum_2_loaded_owners = sum([owners_loaded_count[owner] for owner in ["НПК", "инн"]])
                    sum_2_empty_owners = sum([owners_empty_count[owner] for owner in ["НПК", "инн"]])

                    if owner in ["НПК", "инн"]:

                        if "инн" in ["НПК", "инн"] and "НПК" in ["НПК", "инн"]:
                            owner = "НПК"

                        total = f"{sum_2_loaded_owners}/{sum_2_empty_owners}" if sum_2_loaded_owners > 0 else f"{sum_2_empty_owners}"
                    else:
                        total = f"{loaded_count}/{empty_count}" if loaded_count > 0 else f"{empty_count}"

                    owner_string = f"{total} {owner}"

                    total_nambervag = f"{npk_nambervag}" if owner == 'НПК' else f"{other_nambervag}"
                    owners_strings_set.add(owner_string)
                    unique_owner_strings = list(owners_strings_set)
                    sorted_owner_strings = sorted(unique_owner_strings, key=lambda x: owner_key(x.split()[1]))
                    sorted_owner_strings_not_nambervag = sorted(unique_owner_strings, key=lambda x: owner_key(x.split()[1]))

                    first_owner = sorted_owner_strings[0]
                    if first_owner.split()[1] == 'НПК':
                        total_nambervag = npk_nambervag
                    else:
                        total_nambervag = other_nambervag

                    first_owner_with_nambervag = f"{first_owner} ({total_nambervag})"
                    sorted_owner_strings[0] = first_owner_with_nambervag
                    owners_string = ', '.join(sorted_owner_strings)
                    owners_string_not_nambervag = ', '.join(sorted_owner_strings_not_nambervag)

            # кол-во собственников в каждом поезде (для подсчета в конце)
            owners_count = {"НПК": 0, "инн": 0, "БП": 0, "ТТ": 0, "Нвтр": 0, "КФТ": 0, "ФГК": 0, "ПГК": 0}

            for owner in sorted(owner_counts[index_r], key=owner_key):
                count = owner_counts[index_r][owner]

                if owner in owners_count:
                    owners_count[owner] += count
                sum_first_4_owners = sum([owners_count[owner] for owner in ["НПК", "инн", "БП", "ТТ"]])
                if sum_first_4_owners > 0:
                    group_part = f"({sum_first_4_owners}"
                    other_owners_counts = [owners_count[owner] for owner in ["Нвтр", "КФТ", "ФГК"]]
                    other_owners_string = '+'.join(str(count) for count in other_owners_counts)
                    group_part += f"+{other_owners_string})" if other_owners_string else "+0+0)"
                else:
                    other_owners_counts = [owners_count[owner] for owner in ["Нвтр", "КФТ", "ФГК"]]
                    other_owners_string = '+'.join(str(count) for count in other_owners_counts)
                    group_part = f"(0+{other_owners_string})" if other_owners_string else "(0+0+0)"
                if group_part.endswith('+0+0+0)'):
                    group_part = ''

            # кол-во инновационных вагонов в каждом поезде (в конце)
            owners_count_inn = {"НПК": 0, "инн": 0, "БП": 0, "ТТ": 0, "Нвтр": 0, "КФТ": 0, "ФГК": 0, "ПГК": 0}
            for owner in sorted(owner_counts_inn[index_r], key=owner_key):
                count_inn = owner_counts_inn[index_r][owner]
                if owner in owners_count_inn:
                    owners_count_inn[owner] += count_inn
                npk_inn = [owners_count_inn[owner] for owner in ["инн"]]
                npk_string = ''.join(str(count_inn) for count_inn in npk_inn)
                group_npk = f"({npk_string} инн)" if index_r.endswith('-4384') and npk_string != '0' else ""

                string_train_owner = f"{owners_string_not_nambervag} {group_part} {group_npk}"
                operation_bros = f"{operation} {data_op}" if operation == 'БРОС' else f""
                station_total = f"" if road != '' and not 'БРОС' in operation and station != ' Лихая' and station != 'Батайск' else f"{station}"

                string_total = f"{owners_string} ({index_r}) - {station_total} {road} {operation_bros} {group_part} {group_npk}"

        # запись в лист "брошенные поезда"
        if road == '' and index_r.endswith('-4384') and 'БРОС' in operation:

            worksheet1.write(row1, 0, index_r)
            worksheet1.write(row1, 1, station)
            worksheet1.write(row1, 2, operation)
            worksheet1.write(row1, 3, data_op)
            worksheet1.write(row1, 4, string_train_owner)
            worksheet1.write(row1, 5, total_nambervag)
            row1 += 1
            worksheet1.write(row1, 0, string_bros_train)

        # запись в лист "прямые поезда"
        if road == '' and index_r.endswith('-4384') and not 'БРОС' in operation:
            worksheet2.write(row2, 0, index_r)
            worksheet2.write(row2, 1, station)
            worksheet2.write(row2, 2, operation)
            worksheet2.write(row2, 3, data_op)
            worksheet2.write(row2, 4, string_train_owner)
            worksheet2.write(row2, 5, total_nambervag)
            row2 += 1
            worksheet2.write(row2, 0, string_direct_train)

        # запись в лист "прямые поезда"
        if road != '' and index_r.endswith('-4384') and not 'БРОС' in operation:
            worksheet2.write(row2 + 2, 0, index_r)
            worksheet2.write(row2 + 2, 1, station)
            worksheet2.write(row2 + 2, 2, operation)
            worksheet2.write(row2 + 2, 3, data_op)
            worksheet2.write(row2 + 2, 4, string_train_owner)
            worksheet2.write(row2 + 2, 5, total_nambervag)
            worksheet2.write(row2 + 2, 6, road)
            row2 += 1

        # запись в лист "подход на стойленскую"
        if road == '' and not index_r.endswith('-4384'):
            worksheet.write(row, 0, string_total)
            worksheet.write(row, 1, total_nambervag)
            row += 1

        # запись в лист "подход на стойленскую"
        if road != '' and not index_r.endswith('-4384'):
            worksheet.write(row + 1, 0, string_total)
            worksheet.write(row + 1, 1, total_nambervag)
            row += 1
            worksheet.write(row + 1, 0, string_not_all_train)
            worksheet.write(row + 2, 0, string_all_train)

    workbook.close()

window["ОК"].update(disabled=False)
event, values = window.read()
if event == "ОК" or event == Sg.WINDOW_CLOSED:
    window.close()
