import PySimpleGUI as Sg
from openpyxl import load_workbook
from datetime import datetime
import xlwt

layout = [
        [Sg.Text('Введите количество вагонов в поезде:')],
        [Sg.Input()],
        [Sg.Button('Сформировать отчет')]
]
window = Sg.Window('Подход на ОЭМК', layout)
kolvag = 0
while True:
    event, values = window.read()
    if event == Sg.WINDOW_CLOSED:
        break
    elif event == 'Сформировать отчет':
        kolvag = int(values[0])

    window.close()

layout = [[Sg.Text("Отчет сформирован!")], [Sg.Button("ОК")]]
window = Sg.Window("Формирование отчета...", layout, finalize=True, size=(350, 100))

wb = load_workbook('Подход (ГОКи).xlsx')
sheet = wb.active
max_rows = sheet.max_row

index_dict = {}
index_rail = {}

for i in range(6, max_rows+1):
    nambervag = sheet.cell(row=i, column=2).value

    index_r = sheet.cell(row=i, column=11).value
    if index_r == '0000-000-0000' or index_r.endswith('-5210'):
        continue

    station = sheet.cell(row=i, column=8).value
    station = ' '.join(part.capitalize() for part in station.replace('-', ' ').split())
    if station == 'Старый Оскол':
        station = '   Ст.Оскол'
    elif station == 'Губкин':
        station = '  Губкин'
    elif station == 'Лебеди':
        station = '  Лебели'
    elif station == 'Стойленская':
        station = '  Стойленская'
    elif station == 'Валуйки':
        station = ' Валуйки'
    elif station == 'Котел':
        continue

    index_gp = sheet.cell(row=i, column=16).value
    if index_gp != '5134' and index_gp != '3161':
        continue

    condition = sheet.cell(row=i, column=33).value
    if condition != 'гр' and condition != 'пр':
        continue

    weight = sheet.cell(row=i, column=12).value

    data_op = sheet.cell(row=i, column=9).value
    if isinstance(data_op, datetime):
        data_op = data_op.strftime('%d.%m %H-%M')

    road = sheet.cell(row=i, column=10).value

    if road == 'КБШ':
        road = 'Кбш.ж.д.'
    elif road == 'МСК':
        road = 'Мск.ж.д.'
    elif road == 'ПРВ':
        road = 'Прив.ж.д.'
    elif road == 'СКВ':
        road = 'С-Кав.ж.д.'
    elif road == 'ЮВС':
        road = 'Ю-Вос.ж.д.'
    else:
        continue

    operation = sheet.cell(row=i, column=14).value
    if operation == 'БРОС':
        operation = f'БРОС {data_op}'
    else:
        operation = ''

    cargo = sheet.cell(row=i, column=19).value
    if cargo == 'Глина Бентонит (24122)':
        cargo = 'глина'
    elif cargo == 'Камень Извест (24133)' or cargo == 'Известняк Д/Фл (29103)':
        cargo = 'флюсы'
    elif cargo == 'Лом Чермет Пр (31607)':
        cargo = 'лом'
    elif cargo == 'Руда Флюоритов (24209)':
        cargo = 'шпат'
    elif cargo == 'Электроды Гр Пр (47306)':
        cargo = 'электроды'
    elif cargo == 'Ферросплавы Пр (31308)':
        cargo = 'ферроспл'
    elif cargo == 'УГЛЕРОД ДОБАВКА (24161)':
        cargo = 'уголь'
    elif cargo == 'Брикеты Желез (14103)':
        cargo = 'брикеты'
    elif cargo == 'Окатыши Ж/Руд (14111)':
        cargo = 'окатыши'
    else:
        cargo = " "

    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet('Подход на Котел')

    row = 0
    index_rail.setdefault(road, []).append(weight)
    sorteddict = sorted(index_rail.items(), reverse=True)

    for key, value in sorteddict:
        i = len(value)
        if i < 20:
            continue
        gr_0 = 0
        gr_not_0 = 0
        for weight in value:
            if weight == 0:
                gr_0 += 1
            else:
                gr_not_0 += 1

        string_a = f"{gr_not_0}/{gr_0} пв НПК - {''.join(key)} ({i})"
        worksheet.write(row, 0, string_a)
        row += 1

    row = worksheet.last_used_row + 3
    count_dict = {}
    cargo_gr_dict = {}

    if road != 'Ю-Вос.ж.д.' and operation != f'БРОС {data_op}':
        station = ''
    if road == 'Ю-Вос.ж.д.':
        road = ''

    index_dict.setdefault(index_r, []).append((station, road, operation, condition, cargo, nambervag))
    sorted_values = sorted(index_dict.items(), key=lambda x: (x[1][0][1], x[1][0][0]))
    for index_r, values_list in sorted_values:
        if len(values_list) >= kolvag:
            gr = 0
            pr = 0
            for station, road, operation, condition, cargo, nambervag in values_list:
                if condition != 'пр':
                    gr += 1
                    if cargo != '0':
                        if cargo in cargo_gr_dict:
                            cargo_gr_dict[cargo] += 1
                        else:
                            cargo_gr_dict[cargo] = 1
                else:
                    pr += 1

            if index_r not in count_dict:
                count_dict[index_r] = (gr, pr)
                string_toal = f"{gr}/{pr} пв НПК ({nambervag}) ({index_r}) - {station}  {road}  {operation}"
                worksheet.write(row, 0, [string_toal, ', '.join(
                    [f"({count} {cargo})" if count > 1 else f'({cargo})' for cargo, count in cargo_gr_dict.items()])])
                row += 1
                cargo_gr_dict = {}
                workbook.save("ОЭМК.xls")

window["ОК"].update(disabled=False)
event, values = window.read()
if event == "ОК" or event == Sg.WINDOW_CLOSED:
    window.close()
